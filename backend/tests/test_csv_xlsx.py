"""Tests for csv_xlsx.py.

The batch contract is the same shape as the other converters. What is specific here is
fidelity: pandas infers a dtype per column on the way in, and anything that inference
gets wrong is written into the workbook as the wrong value. The round-trip group at the
bottom pushes a CSV through to XLSX and back and compares.
"""

import shutil
from collections.abc import Callable
from pathlib import Path
from types import ModuleType

import pandas as pd
from openpyxl import load_workbook

import csv_xlsx
import xlsx_csv

Sandbox = Callable[[ModuleType], tuple[Path, Path]]


def round_trip(sandbox: Sandbox, csv_text: str) -> str:
    """Send a CSV through csv_xlsx and back through xlsx_csv, returning the CSV text
    that comes out the far end. Both modules read the same folder globals, so the
    intermediate workbook is moved from output back to input between the two legs."""
    input_dir, output_dir = sandbox(csv_xlsx)
    sandbox(xlsx_csv)

    (input_dir / "data.csv").write_text(csv_text, encoding="utf-8")
    csv_xlsx.convert_csv_to_xlsx()

    (input_dir / "data.csv").unlink()
    shutil.move(output_dir / "data.xlsx", input_dir / "data.xlsx")
    xlsx_csv.convert_xlsx_to_csv()

    return (output_dir / "data.csv").read_text(encoding="utf-8")


# Batch contract


def test_reports_when_input_is_empty(sandbox: Sandbox) -> None:
    sandbox(csv_xlsx)
    assert csv_xlsx.convert_csv_to_xlsx() == "No CSV files found in input folder"


def test_reports_when_the_file_is_already_a_workbook(sandbox: Sandbox) -> None:
    input_dir, _ = sandbox(csv_xlsx)
    pd.DataFrame({"a": [1]}).to_excel(input_dir / "book.xlsx", index=False)

    assert csv_xlsx.convert_csv_to_xlsx() == "That file is already in Excel format"


def test_converts_a_single_csv(sandbox: Sandbox) -> None:
    input_dir, output_dir = sandbox(csv_xlsx)
    (input_dir / "data.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    summary = csv_xlsx.convert_csv_to_xlsx()

    assert "data.xlsx" in summary
    assert pd.read_excel(output_dir / "data.xlsx").to_dict("list") == {"a": [1], "b": [2]}


def test_converts_several_csvs(sandbox: Sandbox) -> None:
    input_dir, output_dir = sandbox(csv_xlsx)
    (input_dir / "one.csv").write_text("a\n1\n", encoding="utf-8")
    (input_dir / "two.csv").write_text("b\n2\n", encoding="utf-8")

    summary = csv_xlsx.convert_csv_to_xlsx()

    assert "Converted 2 file(s)" in summary
    assert sorted(p.name for p in output_dir.iterdir()) == ["one.xlsx", "two.xlsx"]


def test_an_empty_csv_is_reported_rather_than_raised(sandbox: Sandbox) -> None:
    """A zero-byte file has no header row for pandas to read, which raises rather than
    producing an empty frame."""
    input_dir, _ = sandbox(csv_xlsx)
    (input_dir / "empty.csv").write_text("", encoding="utf-8")

    summary = csv_xlsx.convert_csv_to_xlsx()

    assert summary.startswith("No files converted. 1 failed:")
    assert "empty.csv" in summary


def test_one_bad_file_does_not_stop_the_good_ones(sandbox: Sandbox) -> None:
    input_dir, output_dir = sandbox(csv_xlsx)
    (input_dir / "good.csv").write_text("a\n1\n", encoding="utf-8")
    (input_dir / "empty.csv").write_text("", encoding="utf-8")

    summary = csv_xlsx.convert_csv_to_xlsx()

    assert "good.xlsx" in summary
    assert "1 failed" in summary
    assert (output_dir / "good.xlsx").exists()


# Round-trip fidelity


def test_text_survives_the_round_trip(sandbox: Sandbox) -> None:
    assert round_trip(sandbox, "word\ncafé\n日本語\n") == "word\ncafé\n日本語\n"


def test_a_quoted_comma_survives_the_round_trip(sandbox: Sandbox) -> None:
    """The comma inside the value must stay inside one field rather than splitting it."""
    source = 'name\n"Smith, John"\n'
    assert round_trip(sandbox, source) == source


def test_an_empty_cell_survives_the_round_trip(sandbox: Sandbox) -> None:
    assert round_trip(sandbox, "a,b\n1,\n") == "a,b\n1,\n"


def test_a_date_like_string_survives_the_round_trip(sandbox: Sandbox) -> None:
    assert round_trip(sandbox, "d\n2024-01-02\n") == "d\n2024-01-02\n"


def test_leading_zeros_survive_the_round_trip(sandbox: Sandbox) -> None:
    """A zip code, product code or phone number written as 007 is a string, but pandas
    infers the column as integer and the zeros are gone for good."""
    assert round_trip(sandbox, "code\n007\n042\n") == "code\n007\n042\n"


def test_boolean_spelling_survives_the_round_trip(sandbox: Sandbox) -> None:
    """TRUE is inferred as a Python bool and comes back spelled True."""
    assert round_trip(sandbox, "flag\nTRUE\nFALSE\n") == "flag\nTRUE\nFALSE\n"


def test_a_cell_starting_with_equals_is_not_turned_into_a_formula(
    sandbox: Sandbox,
) -> None:
    """openpyxl stores a leading = as a live formula rather than as the text it was.
    The workbook has no cached result for it, so the value does not survive, and a
    spreadsheet opening that file would execute it."""
    assert round_trip(sandbox, "f\n=1+1\n") == "f\n=1+1\n"


def test_trailing_zeros_on_a_float_survive_the_round_trip(sandbox: Sandbox) -> None:
    """1.10 and 1.1 are the same number but not the same text, so the column is kept as
    text rather than converted. A price list written to two decimal places stays that
    way instead of being reformatted on the user's behalf."""
    assert round_trip(sandbox, "n\n1.10\n") == "n\n1.10\n"


def test_a_plain_numeric_column_is_still_written_as_numbers(sandbox: Sandbox) -> None:
    """The point of detecting per column rather than reading everything as text: a
    column that loses nothing by being numeric stays numeric, so Excel can still add
    it up."""
    input_dir, output_dir = sandbox(csv_xlsx)
    (input_dir / "data.csv").write_text("price,code\n15,007\n20,042\n", encoding="utf-8")

    csv_xlsx.convert_csv_to_xlsx()

    # Read the cells rather than a DataFrame: what matters is the type stored in the
    # workbook, and pandas would re-guess it on the way back in.
    sheet = load_workbook(output_dir / "data.xlsx").active
    price = [sheet.cell(row=r, column=1) for r in (2, 3)]
    code = [sheet.cell(row=r, column=2) for r in (2, 3)]

    assert [c.data_type for c in price] == ["n", "n"]
    assert [c.value for c in price] == [15, 20]
    assert [c.data_type for c in code] == ["s", "s"]
    assert [c.value for c in code] == ["007", "042"]
